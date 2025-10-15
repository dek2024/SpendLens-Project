"""
Logic module for SpendLens - Voice Expense Logger
Contains helper functions for data parsing, manipulation, and processing.
"""

import re
import pandas as pd
from datetime import datetime, timedelta
from typing import Optional, Tuple
from word2number import w2n
import dateparser
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference


def extract_amount(text: str) -> float:
    """
    Extract numeric or written dollar amounts from text.
    
    Args:
        text: Input text containing potential amount information
        
    Returns:
        Extracted amount as float, or 0.0 if no amount found
        
    Examples:
        >>> extract_amount("I spent $15.50 today")
        15.5
        >>> extract_amount("Cost was twenty dollars")
        20.0
    """
    if not text:
        return 0.0
    
    text_clean = text.replace(",", "").replace("$", "").lower()
    
    # Try to find numeric amount first
    match = re.search(r"\b(\d+(?:\.\d{1,2})?)\b", text_clean)
    if match:
        try:
            return float(match.group(1))
        except ValueError:
            pass
    
    # Try to convert written numbers to numeric
    try:
        return float(w2n.word_to_num(text_clean))
    except:
        return 0.0


def extract_date(text: str) -> datetime.date:
    """
    Detect and parse date phrases from text.
    
    Handles natural language date expressions like:
    - 'yesterday', 'tomorrow', 'today'
    - 'next Wednesday', 'last Friday'
    - '3 days ago', 'two weeks ago'
    
    Args:
        text: Input text containing potential date information
        
    Returns:
        Parsed date as datetime.date object, defaults to today if no date found
        
    Examples:
        >>> extract_date("I spent money yesterday")
        # Returns yesterday's date
        >>> extract_date("next Monday")
        # Returns date of next Monday
    """
    if not text:
        return datetime.now().date()

    text_lower = text.lower()
    today = datetime.now()

    # Handle common explicit date phrases
    if "yesterday" in text_lower:
        return (today - timedelta(days=1)).date()
    elif "tomorrow" in text_lower:
        return (today + timedelta(days=1)).date()
    elif "last" in text_lower or "ago" in text_lower:
        parsed_date = dateparser.parse(
            text, 
            settings={'PREFER_DATES_FROM': 'past', 'RELATIVE_BASE': today}
        )
        if parsed_date:
            return parsed_date.date()
        return today.date()
    elif "next" in text_lower:
        parsed_date = dateparser.parse(
            text, 
            settings={'PREFER_DATES_FROM': 'future', 'RELATIVE_BASE': today}
        )
        if parsed_date:
            return parsed_date.date()
        return today.date()

    # Try general date parsing
    parsed_date = dateparser.parse(text, settings={'RELATIVE_BASE': today})
    if parsed_date:
        return parsed_date.date()

    return today.date()


def load_expense_data(file_path: str) -> pd.DataFrame:
    """
    Load expense data from Excel file or create new DataFrame.
    
    Args:
        file_path: Path to the Excel file
        
    Returns:
        DataFrame with columns: Date/Time, Category, Amount ($), Notes
    """
    import os
    
    if os.path.exists(file_path):
        return pd.read_excel(file_path)
    else:
        return pd.DataFrame(columns=["Date/Time", "Category", "Amount ($)", "Notes"])


def save_expense_data(df: pd.DataFrame, file_path: str) -> None:
    """
    Save expense DataFrame to Excel file.
    
    Args:
        df: DataFrame containing expense data
        file_path: Path where the Excel file should be saved
    """
    df.to_excel(file_path, index=False)


def add_expense_entry(
    df: pd.DataFrame,
    date: datetime.date,
    category: str,
    amount: float,
    note: str
) -> pd.DataFrame:
    """
    Add a new expense entry to the DataFrame.
    
    Args:
        df: Existing expense DataFrame
        date: Date of the expense
        category: Category of the expense
        amount: Amount spent in dollars
        note: Additional notes about the expense
        
    Returns:
        Updated DataFrame with the new entry added
    """
    new_row = pd.DataFrame({
        "Date/Time": [date.strftime("%Y-%m-%d")],
        "Category": [category],
        "Amount ($)": [amount],
        "Notes": [note]
    })
    return pd.concat([df, new_row], ignore_index=True)


def filter_valid_expenses(df: pd.DataFrame) -> pd.DataFrame:
    """
    Filter out invalid or summary rows from expense data.
    
    Removes rows where:
    - Category is empty/null
    - Category is 'TOTAL' (summary row)
    
    Args:
        df: Raw expense DataFrame
        
    Returns:
        Filtered DataFrame containing only valid expense entries
    """
    return df[
        (df["Category"].notna()) &
        (df["Category"].str.upper() != "TOTAL")
    ]


def calculate_category_totals(df: pd.DataFrame) -> pd.Series:
    """
    Calculate total spending by category.
    
    Args:
        df: Expense DataFrame (should be pre-filtered with filter_valid_expenses)
        
    Returns:
        Series with categories as index and total amounts as values, sorted descending
    """
    return df.groupby("Category")["Amount ($)"].sum().sort_values(ascending=False)


def add_total_row(df: pd.DataFrame) -> pd.DataFrame:
    """
    Add a TOTAL summary row to the expense DataFrame.
    
    Args:
        df: Expense DataFrame
        
    Returns:
        DataFrame with TOTAL row appended
    """
    total_spent = df["Amount ($)"].sum()
    total_row = pd.DataFrame({
        "Date/Time": [""],
        "Category": ["TOTAL"],
        "Amount ($)": [total_spent],
        "Notes": [""]
    })
    return pd.concat([df, total_row], ignore_index=True)


def format_excel_export(file_path: str) -> None:
    """
    Apply professional formatting to Excel export file.
    
    Applies:
    - Styled headers with dark blue background
    - Alternating row colors for readability
    - Currency formatting for amounts
    - Borders and cell alignment
    - Auto-adjusted column widths
    - Spending by Category bar chart
    
    Args:
        file_path: Path to the Excel file to format
    """
    wb = load_workbook(file_path)
    ws = wb.active
    ws.title = "Expense Summary"

    # Header styling
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Alternating row colors
    alt_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    for row in range(2, ws.max_row + 1):
        if row % 2 == 0:
            for cell in ws[row]:
                cell.fill = alt_fill

    # Currency formatting for Amount column
    for cell in ws["C"][1:]:
        cell.number_format = '"$"#,##0.00'

    # Borders and alignment
    thin_border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC")
    )
    for row in ws.iter_rows():
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    # Auto-adjust column widths
    for col_cells in ws.columns:
        max_len = max(len(str(c.value)) if c.value else 0 for c in col_cells)
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = max_len + 4

    # Add bar chart
    chart = BarChart()
    chart.title = "Spending by Category"
    chart.y_axis.title = "Amount ($)"
    chart.x_axis.title = "Category"

    data = Reference(ws, min_col=3, min_row=1, max_row=ws.max_row - 1)
    categories = Reference(ws, min_col=2, min_row=2, max_row=ws.max_row - 1)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    ws.add_chart(chart, f"A{ws.max_row + 3}")

    wb.save(file_path)


def prepare_export_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    """
    Prepare expense DataFrame for export by filtering and adding total row.
    
    Args:
        df: Raw expense DataFrame
        
    Returns:
        Export-ready DataFrame with valid expenses and TOTAL row
    """
    export_df = filter_valid_expenses(df).copy()
    if not export_df.empty:
        export_df = add_total_row(export_df)
    return export_df
