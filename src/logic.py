"""
Business logic for SpendLens application.
Implements SOLID principles with clean OOP architecture.

Design Patterns:
- Single Responsibility: Each class has one clear purpose
- Open/Closed: Extensible via interfaces
- Liskov Substitution: All implementations are interchangeable
- Interface Segregation: Focused interfaces (IStorage, IParser, IAIService)
- Dependency Inversion: Depends on abstractions, not concretions
"""

import re
import logging
from abc import ABC, abstractmethod
from datetime import datetime, timedelta
from typing import List, Dict
import pandas as pd
from word2number import w2n
import dateparser
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.chart import BarChart, Reference

from models import Expense, CategoryTotal, ParsedExpense

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s"
)
logger = logging.getLogger(__name__)


# ==================== INTERFACES ====================

class IParser(ABC):
    @abstractmethod
    def parse_amount(self, text: str) -> float:
        pass

    @abstractmethod
    def parse_date(self, text: str) -> datetime:
        pass

    @abstractmethod
    def parse_expense(self, text: str) -> ParsedExpense:
        pass


class IStorage(ABC):
    @abstractmethod
    def load_expenses(self) -> List[Expense]:
        pass

    @abstractmethod
    def save_expenses(self, expenses: List[Expense]) -> None:
        pass

    @abstractmethod
    def add_expense(self, expense: Expense) -> None:
        pass

    @abstractmethod
    def clear_all(self) -> None:
        pass


class IAIService(ABC):
    @abstractmethod
    def transcribe_audio(self, audio_file) -> str:
        pass

    @abstractmethod
    def analyze_expenses(self, expenses: List[Expense], query: str) -> str:
        pass


# ==================== PARSER ====================

class ExpenseParser(IParser):
    """Enhanced parser with smarter numeric + word-based amount detection."""

    def parse_amount(self, text: str) -> float:
        """Extract numeric or written dollar amounts from text, including commas and large numbers."""
        if not text:
            logger.error("Failed to parse amount from empty input.")
            return 0.0

        text_clean = text.lower().replace(",", "").strip()
        try:
            # Handle direct $amount early — supports commas and decimals
            dollar_sign_match = re.search(r"\$\s*([\d,]+(?:\.\d{1,2})?)", text)
            if dollar_sign_match:
                value = float(dollar_sign_match.group(1).replace(",", ""))
                logger.info("Parsed numeric amount successfully.")
                return value

            # Numeric fallback (handles large numbers and commas)
            match = re.search(r"\b(\d{1,3}(?:,\d{3})*(?:\.\d{1,2})?|\d+(?:\.\d{1,2})?)\b", text)
            if match:
                value = float(match.group(1).replace(",", ""))
                logger.info("Parsed numeric amount successfully.")
                return value

            # Written numbers (e.g., "twenty five dollars")
            word_match = re.search(r"([\w\s-]+)\s*(?:dollar|bucks?)", text_clean)
            if word_match:
                words = word_match.group(1).replace("-", " ").strip()
                value = float(w2n.word_to_num(words))
                logger.info("Parsed numeric amount successfully.")
                return value

            return 0.0

        except Exception as e:
            logger.error(f"Failed to parse amount from '{text}': {e}")
            return 0.0

    def parse_date(self, text: str) -> datetime:
        if not text:
            return datetime.now()
        text_lower = text.lower()
        today = datetime.now()
        if "yesterday" in text_lower:
            return today - timedelta(days=1)
        if "tomorrow" in text_lower:
            return today + timedelta(days=1)
        parsed = dateparser.parse(text, settings={'RELATIVE_BASE': today})
        return parsed or datetime.now()

    def parse_expense(self, text: str) -> ParsedExpense:
        amount = self.parse_amount(text)
        date = self.parse_date(text)
        return ParsedExpense(text, amount, date, 1.0 if amount > 0 else 0.5)


# ==================== STORAGE ====================

class ExcelStorage(IStorage):
    def __init__(self, file_path: str):
        self.file_path = file_path

    def load_expenses(self) -> List[Expense]:
        import os
        if not os.path.exists(self.file_path):
            return []
        try:
            df = pd.read_excel(self.file_path)
            if df.empty or "Category" not in df.columns:
                logger.warning("Excel file missing headers or empty.")
                return []
            df = df[(df["Category"].notna()) & (df["Category"].str.upper() != "TOTAL")]
            return [Expense.from_dict(row) for _, row in df.iterrows()]
        except Exception as e:
            logger.error(f"Failed to load expenses: {e}")
            return []

    def save_expenses(self, expenses: List[Expense]) -> None:
        try:
            df = pd.DataFrame([
                {
                    "Date/Time": e.date.strftime("%Y-%m-%d"),
                    "Category": e.category,
                    "Amount ($)": e.amount,
                    "Notes": e.notes
                }
                for e in expenses
            ])
            if df.empty:
                df = pd.DataFrame(columns=["Date/Time", "Category", "Amount ($)", "Notes"])
            df.to_excel(self.file_path, index=False)
            logger.info("Saved expenses to Excel file.")
        except Exception as e:
            logger.error(f"Failed to save expenses: {e}")
            raise

    def add_expense(self, expense: Expense) -> None:
        expenses = self.load_expenses()
        expenses.append(expense)
        self.save_expenses(expenses)

    def clear_all(self) -> None:
        """Clear all expenses but preserve headers."""
        try:
            df = pd.DataFrame(columns=["Date/Time", "Category", "Amount ($)", "Notes"])
            df.to_excel(self.file_path, index=False)
            logger.warning("All expense data cleared. File reset with headers.")
        except Exception as e:
            logger.error(f"Failed to clear all expenses: {e}")
            raise


# ==================== ANALYZER ====================

class ExpenseAnalyzer:
    def calculate_category_totals(self, expenses: List[Expense]) -> List[CategoryTotal]:
        category_data: Dict[str, Dict] = {}
        for exp in expenses:
            category_data.setdefault(exp.category, {"total": 0.0, "count": 0})
            category_data[exp.category]["total"] += exp.amount
            category_data[exp.category]["count"] += 1
        totals = [
            CategoryTotal(cat, data["total"], data["count"])
            for cat, data in category_data.items()
        ]
        totals.sort(key=lambda x: x.total, reverse=True)
        return totals

    def get_total_spending(self, expenses: List[Expense]) -> float:
        return sum(exp.amount for exp in expenses)

    def filter_by_date_range(self, expenses: List[Expense], start: datetime, end: datetime) -> List[Expense]:
        """Filter expenses between start and end dates inclusive."""
        return [exp for exp in expenses if start <= exp.date <= end]


# ==================== EXCEL FORMATTER ====================

class ExcelFormatter:
    """Formats Excel files with professional styling, filters, and charts."""

    def format_workbook(self, file_path: str) -> None:
        try:
            df = pd.read_excel(file_path)
            if df.empty:
                raise ValueError("No data found to format.")

            wb = Workbook()
            ws = wb.active
            ws.title = "Expenses"

            # Write headers and data
            for c_idx, col_name in enumerate(df.columns, 1):
                ws.cell(row=1, column=c_idx, value=col_name)
            for r_idx, row in enumerate(df.itertuples(index=False), 2):
                for c_idx, value in enumerate(row, 1):
                    ws.cell(row=r_idx, column=c_idx, value=value)

            # Header styling
            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")

            # Alternating row colors + currency
            for row in range(2, ws.max_row + 1):
                fill_color = "F2F2F2" if row % 2 == 0 else "FFFFFF"
                for cell in ws[row]:
                    cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                    if cell.column == 3:
                        cell.number_format = '"$"#,##0.00'

            # Borders
            thin = Side(border_style="thin", color="999999")
            for row in ws.iter_rows():
                for cell in row:
                    cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                    cell.alignment = Alignment(vertical="center", wrap_text=True)

            # Auto column widths
            for col in ws.columns:
                max_len = max(len(str(c.value)) if c.value else 0 for c in col)
                ws.column_dimensions[get_column_letter(col[0].column)].width = max_len + 3

            # Add table
            table_ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
            table = Table(displayName="ExpenseTable", ref=table_ref)
            style = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
            table.tableStyleInfo = style
            ws.add_table(table)

            # Add bar chart
            chart = BarChart()
            chart.title = "Spending by Category"
            chart.y_axis.title = "Amount ($)"
            chart.x_axis.title = "Category"

            if "Category" in df.columns and "Amount ($)" in df.columns:
                cats = Reference(ws, min_col=2, min_row=2, max_row=ws.max_row)
                vals = Reference(ws, min_col=3, min_row=1, max_row=ws.max_row)
                chart.add_data(vals, titles_from_data=True)
                chart.set_categories(cats)
                ws.add_chart(chart, "E2")

            wb.save(file_path)
            logger.info(f"Excel export formatted successfully: {file_path}")
        except Exception as e:
            logger.error(f"Failed to format Excel file: {e}")
            raise


# ==================== AI + CONTROLLER ====================

class OpenAIService(IAIService):
    def __init__(self, client):
        self.client = client

    def transcribe_audio(self, audio_file) -> str:
        try:
            transcript = self.client.audio.transcriptions.create(
                model="gpt-4o-mini-transcribe",
                file=audio_file
            )
            return transcript.text.strip()
        except Exception as e:
            logger.error(f"Transcription failed: {e}")
            raise

    def analyze_expenses(self, expenses: List[Expense], query: str) -> str:
        try:
            expense_data = "\n".join([
                f"{exp.date.strftime('%Y-%m-%d')}: ${exp.amount:.2f} - {exp.category} ({exp.notes})"
                for exp in expenses
            ])
            response = self.client.chat.completions.create(
                model="gpt-4o-mini",
                messages=[
                    {"role": "system", "content": "You are a financial assistant analyzing expense data."},
                    {"role": "user", "content": f"Here are my expenses:\n{expense_data}\n\nQuestion: {query}"},
                ],
            )
            return response.choices[0].message.content
        except Exception as e:
            logger.error(f"Expense analysis failed: {e}")
            return f"Sorry, I couldn't analyze your expenses: {e}"


# ==================== CONTROLLER ====================

class ExpenseController:
    def __init__(self, storage, parser, analyzer, formatter, ai_service=None):
        self.storage = storage
        self.parser = parser
        self.analyzer = analyzer
        self.formatter = formatter
        self.ai_service = ai_service
        logger.info("ExpenseController initialized.")

    def parse_and_create_expense(self, text, category, manual_amount=None, manual_date=None):
        parsed = self.parser.parse_expense(text)
        return Expense(
            date=manual_date or parsed.detected_date,
            category=category,
            amount=manual_amount if manual_amount is not None else parsed.detected_amount,
            notes=text,
        )

    def add_expense(self, text, category, manual_amount=None, manual_date=None):
        expense = self.parse_and_create_expense(text, category, manual_amount, manual_date)
        self.storage.add_expense(expense)
        return expense

    def get_all_expenses(self):
        """Return all expenses currently stored."""
        return self.storage.load_expenses()

    def get_dashboard_data(self):
        expenses = self.storage.load_expenses()
        category_totals = self.analyzer.calculate_category_totals(expenses)
        total_spending = self.analyzer.get_total_spending(expenses)
        return {
            "expenses": expenses,
            "category_totals": category_totals,
            "total_spending": total_spending,
            "expense_count": len(expenses),
        }

    def export_to_excel(self, file_path):
        expenses = self.storage.load_expenses()
        if not expenses:
            logger.warning("No expenses to export.")
            df = pd.DataFrame(columns=["Date/Time", "Category", "Amount ($)", "Notes"])
            df.to_excel(file_path, index=False)
            return
        total = self.analyzer.get_total_spending(expenses)
        expenses.append(Expense(datetime.now(), "TOTAL", total, ""))
        ExcelStorage(file_path).save_expenses(expenses)
        self.formatter.format_workbook(file_path)

    def clear_all_expenses(self):
        self.storage.clear_all()
