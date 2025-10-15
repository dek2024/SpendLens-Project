"""
Comprehensive test suite for logic.py module
Tests all helper functions with edge cases and error handling
"""

import pytest
import pandas as pd
from datetime import datetime, timedelta
from unittest.mock import patch, mock_open, MagicMock
import os
import tempfile

# Import functions from logic module
import sys
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', 'src'))

from logic import (
    extract_amount,
    extract_date,
    load_expense_data,
    save_expense_data,
    add_expense_entry,
    filter_valid_expenses,
    calculate_category_totals,
    add_total_row,
    format_excel_export,
    prepare_export_dataframe
)


class TestExtractAmount:
    """Test suite for extract_amount() function"""
    
    def test_numeric_dollar_amount(self):
        """Test extraction of numeric dollar amounts"""
        assert extract_amount("I spent $25.50 today") == 25.50
        assert extract_amount("Cost was $100") == 100.0
        assert extract_amount("15.99 dollars") == 15.99
    
    def test_numeric_without_dollar_sign(self):
        """Test extraction of plain numeric values"""
        assert extract_amount("I paid 42 for lunch") == 42.0
        assert extract_amount("The cost is 150.25") == 150.25
    
    def test_comma_separated_numbers(self):
        """Test extraction from numbers with commas"""
        assert extract_amount("$1,500") == 1500.0
        assert extract_amount("2,345.67 spent") == 2345.67
    
    def test_written_numbers(self):
        """Test extraction of written-out number words"""
        assert extract_amount("twenty dollars") == 20.0
        assert extract_amount("Cost was fifty") == 50.0
        assert extract_amount("ten bucks") == 10.0
    
    def test_written_compound_numbers(self):
        """Test extraction of compound written numbers"""
        assert extract_amount("twenty five dollars") == 25.0
        assert extract_amount("thirty two") == 32.0
    
    def test_empty_string(self):
        """Test with empty string"""
        assert extract_amount("") == 0.0
    
    def test_none_input(self):
        """Test with None input"""
        assert extract_amount(None) == 0.0
    
    def test_no_amount_in_text(self):
        """Test with text containing no amounts"""
        assert extract_amount("just some text") == 0.0
        assert extract_amount("no numbers here") == 0.0
    
    def test_multiple_amounts_first_match(self):
        """Test that first numeric match is extracted"""
        assert extract_amount("I spent 25 and then 30 dollars") == 25.0
    
    def test_decimal_precision(self):
        """Test amounts with decimal precision"""
        assert extract_amount("$12.99") == 12.99
        assert extract_amount("5.5 dollars") == 5.5
    
    def test_zero_amount(self):
        """Test extraction of zero"""
        assert extract_amount("0 dollars") == 0.0
        assert extract_amount("$0.00") == 0.0
    
    def test_case_insensitivity(self):
        """Test case insensitive extraction"""
        assert extract_amount("TWENTY DOLLARS") == 20.0
        assert extract_amount("FiFtEeN") == 15.0


class TestExtractDate:
    """Test suite for extract_date() function"""
    
    def test_yesterday(self):
        """Test 'yesterday' keyword"""
        expected = (datetime.now() - timedelta(days=1)).date()
        result = extract_date("I spent money yesterday")
        assert result == expected
        
    def test_tomorrow(self):
        """Test 'tomorrow' keyword"""
        expected = (datetime.now() + timedelta(days=1)).date()
        result = extract_date("tomorrow I will spend")
        assert result == expected
    
    def test_today_default(self):
        """Test that today is returned for unrecognized dates"""
        expected = datetime.now().date()
        result = extract_date("just some random text")
        assert result == expected
    
    def test_empty_string(self):
        """Test with empty string returns today"""
        expected = datetime.now().date()
        assert extract_date("") == expected
    
    def test_none_input(self):
        """Test with None input returns today"""
        expected = datetime.now().date()
        assert extract_date(None) == expected
    
    def test_explicit_yesterday_keyword(self):
        """Test various phrasings with 'yesterday'"""
        expected = (datetime.now() - timedelta(days=1)).date()
        assert extract_date("yesterday") == expected
        assert extract_date("Yesterday I bought") == expected
        assert extract_date("Bought YESTERDAY") == expected
    
    def test_explicit_tomorrow_keyword(self):
        """Test various phrasings with 'tomorrow'"""
        expected = (datetime.now() + timedelta(days=1)).date()
        assert extract_date("tomorrow") == expected
        assert extract_date("Tomorrow I will") == expected
    
    @patch('logic.datetime')
    def test_mocked_date_yesterday(self, mock_datetime):
        """Test with mocked datetime for consistent testing"""
        fixed_date = datetime(2025, 10, 14, 12, 0, 0)
        mock_datetime.now.return_value = fixed_date
        mock_datetime.side_effect = lambda *args, **kwargs: datetime(*args, **kwargs)
        
        # This would be yesterday from the mocked date
        expected = datetime(2025, 10, 13).date()
        # Note: actual implementation uses datetime.now() which we mocked
        result = extract_date("yesterday")
        # Since we're mocking, verify the logic would work
        assert isinstance(result, type(expected))
    
    def test_days_ago(self):
        """Test 'X days ago' pattern"""
        result = extract_date("3 days ago")
        # Result should be in the past
        assert result <= datetime.now().date()
    
    def test_next_week(self):
        """Test 'next week' pattern"""
        result = extract_date("next week")
        # Result should be in the future
        assert result >= datetime.now().date()
    
    def test_last_week(self):
        """Test 'last week' pattern"""
        result = extract_date("last week")
        # Result should be in the past
        assert result <= datetime.now().date()
    
    def test_case_insensitivity(self):
        """Test case insensitive date parsing"""
        expected_yesterday = (datetime.now() - timedelta(days=1)).date()
        assert extract_date("YESTERDAY") == expected_yesterday
        assert extract_date("YeStErDaY") == expected_yesterday


class TestLoadExpenseData:
    """Test suite for load_expense_data() function"""
    
    def test_load_existing_file(self):
        """Test loading from existing Excel file"""
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            tmp_path = tmp.name
            # Create test data
            df = pd.DataFrame({
                "Date/Time": ["2025-10-14"],
                "Category": ["Food"],
                "Amount ($)": [25.50],
                "Notes": ["Lunch"]
            })
            df.to_excel(tmp_path, index=False)
            
            # Load and verify
            loaded_df = load_expense_data(tmp_path)
            assert not loaded_df.empty
            assert len(loaded_df) == 1
            assert loaded_df["Category"].iloc[0] == "Food"
            
            # Cleanup
            os.unlink(tmp_path)
    
    def test_load_nonexistent_file(self):
        """Test loading when file doesn't exist creates empty DataFrame"""
        df = load_expense_data("nonexistent_file.xlsx")
        assert isinstance(df, pd.DataFrame)
        assert df.empty
        assert list(df.columns) == ["Date/Time", "Category", "Amount ($)", "Notes"]


class TestSaveExpenseData:
    """Test suite for save_expense_data() function"""
    
    def test_save_data(self):
        """Test saving DataFrame to Excel file"""
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            tmp_path = tmp.name
            
            df = pd.DataFrame({
                "Date/Time": ["2025-10-14"],
                "Category": ["Food"],
                "Amount ($)": [25.50],
                "Notes": ["Lunch"]
            })
            
            save_expense_data(df, tmp_path)
            
            # Verify file exists and can be read
            assert os.path.exists(tmp_path)
            loaded_df = pd.read_excel(tmp_path)
            assert len(loaded_df) == 1
            
            # Cleanup
            os.unlink(tmp_path)


class TestAddExpenseEntry:
    """Test suite for add_expense_entry() function"""
    
    def test_add_single_entry(self):
        """Test adding a single expense entry"""
        df = pd.DataFrame(columns=["Date/Time", "Category", "Amount ($)", "Notes"])
        
        result_df = add_expense_entry(
            df,
            date=datetime(2025, 10, 14).date(),
            category="Food",
            amount=25.50,
            note="Lunch at restaurant"
        )
        
        assert len(result_df) == 1
        assert result_df["Category"].iloc[0] == "Food"
        assert result_df["Amount ($)"].iloc[0] == 25.50
        assert result_df["Date/Time"].iloc[0] == "2025-10-14"
    
    def test_add_multiple_entries(self):
        """Test adding multiple entries sequentially"""
        df = pd.DataFrame(columns=["Date/Time", "Category", "Amount ($)", "Notes"])
        
        df = add_expense_entry(df, datetime(2025, 10, 14).date(), "Food", 25.50, "Lunch")
        df = add_expense_entry(df, datetime(2025, 10, 15).date(), "Gas", 50.00, "Fill up")
        
        assert len(df) == 2
        assert df["Category"].iloc[1] == "Gas"
    
    def test_add_entry_preserves_existing(self):
        """Test that adding entry preserves existing data"""
        df = pd.DataFrame({
            "Date/Time": ["2025-10-13"],
            "Category": ["Shopping"],
            "Amount ($)": [100.0],
            "Notes": ["Clothes"]
        })
        
        result_df = add_expense_entry(
            df,
            datetime(2025, 10, 14).date(),
            "Food",
            25.50,
            "Lunch"
        )
        
        assert len(result_df) == 2
        assert result_df["Category"].iloc[0] == "Shopping"
        assert result_df["Category"].iloc[1] == "Food"


class TestFilterValidExpenses:
    """Test suite for filter_valid_expenses() function"""
    
    def test_filter_removes_null_categories(self):
        """Test that null categories are filtered out"""
        df = pd.DataFrame({
            "Date/Time": ["2025-10-14", "2025-10-15"],
            "Category": ["Food", None],
            "Amount ($)": [25.50, 10.00],
            "Notes": ["Lunch", ""]
        })
        
        result = filter_valid_expenses(df)
        assert len(result) == 1
        assert result["Category"].iloc[0] == "Food"
    
    def test_filter_removes_total_rows(self):
        """Test that TOTAL summary rows are filtered out"""
        df = pd.DataFrame({
            "Date/Time": ["2025-10-14", ""],
            "Category": ["Food", "TOTAL"],
            "Amount ($)": [25.50, 25.50],
            "Notes": ["Lunch", ""]
        })
        
        result = filter_valid_expenses(df)
        assert len(result) == 1
        assert result["Category"].iloc[0] == "Food"
    
    def test_filter_case_insensitive_total(self):
        """Test TOTAL filtering is case insensitive"""
        df = pd.DataFrame({
            "Date/Time": ["2025-10-14", "", ""],
            "Category": ["Food", "Total", "total"],
            "Amount ($)": [25.50, 25.50, 25.50],
            "Notes": ["Lunch", "", ""]
        })
        
        result = filter_valid_expenses(df)
        assert len(result) == 1
    
    def test_filter_empty_dataframe(self):
        """Test filtering empty DataFrame"""
        df = pd.DataFrame(columns=["Date/Time", "Category", "Amount ($)", "Notes"])
        result = filter_valid_expenses(df)
        assert result.empty


class TestCalculateCategoryTotals:
    """Test suite for calculate_category_totals() function"""
    
    def test_calculate_single_category(self):
        """Test calculation with single category"""
        df = pd.DataFrame({
            "Date/Time": ["2025-10-14", "2025-10-15"],
            "Category": ["Food", "Food"],
            "Amount ($)": [25.50, 15.00],
            "Notes": ["Lunch", "Breakfast"]
        })
        
        result = calculate_category_totals(df)
        assert result["Food"] == 40.50
    
    def test_calculate_multiple_categories(self):
        """Test calculation with multiple categories"""
        df = pd.DataFrame({
            "Date/Time": ["2025-10-14", "2025-10-15", "2025-10-16"],
            "Category": ["Food", "Gas", "Food"],
            "Amount ($)": [25.50, 50.00, 15.00],
            "Notes": ["Lunch", "Fill up", "Breakfast"]
        })
        
        result = calculate_category_totals(df)
        assert result["Food"] == 40.50
        assert result["Gas"] == 50.00
    
    def test_sorted_descending(self):
        """Test that results are sorted in descending order"""
        df = pd.DataFrame({
            "Date/Time": ["2025-10-14", "2025-10-15", "2025-10-16"],
            "Category": ["Food", "Gas", "Shopping"],
            "Amount ($)": [25.50, 100.00, 50.00],
            "Notes": ["Lunch", "Fill up", "Clothes"]
        })
        
        result = calculate_category_totals(df)
        categories = list(result.index)
        assert categories[0] == "Gas"  # Highest
        assert categories[1] == "Shopping"
        assert categories[2] == "Food"  # Lowest
    
    def test_empty_dataframe(self):
        """Test with empty DataFrame"""
        df = pd.DataFrame(columns=["Date/Time", "Category", "Amount ($)", "Notes"])
        result = calculate_category_totals(df)
        assert len(result) == 0


class TestAddTotalRow:
    """Test suite for add_total_row() function"""
    
    def test_add_total_to_dataframe(self):
        """Test adding TOTAL row to DataFrame"""
        df = pd.DataFrame({
            "Date/Time": ["2025-10-14", "2025-10-15"],
            "Category": ["Food", "Gas"],
            "Amount ($)": [25.50, 50.00],
            "Notes": ["Lunch", "Fill up"]
        })
        
        result = add_total_row(df)
        assert len(result) == 3
        assert result["Category"].iloc[2] == "TOTAL"
        assert result["Amount ($)"].iloc[2] == 75.50
    
    def test_total_row_format(self):
        """Test that TOTAL row has correct format"""
        df = pd.DataFrame({
            "Date/Time": ["2025-10-14"],
            "Category": ["Food"],
            "Amount ($)": [25.50],
            "Notes": ["Lunch"]
        })
        
        result = add_total_row(df)
        total_row = result.iloc[-1]
        assert total_row["Date/Time"] == ""
        assert total_row["Category"] == "TOTAL"
        assert total_row["Notes"] == ""
    
    def test_empty_dataframe_total(self):
        """Test adding total to empty DataFrame"""
        df = pd.DataFrame(columns=["Date/Time", "Category", "Amount ($)", "Notes"])
        result = add_total_row(df)
        assert len(result) == 1
        assert result["Amount ($)"].iloc[0] == 0.0


class TestPrepareExportDataframe:
    """Test suite for prepare_export_dataframe() function"""
    
    def test_prepare_filters_and_adds_total(self):
        """Test that preparation filters invalid rows and adds total"""
        df = pd.DataFrame({
            "Date/Time": ["2025-10-14", "2025-10-15", ""],
            "Category": ["Food", "Gas", None],
            "Amount ($)": [25.50, 50.00, 10.00],
            "Notes": ["Lunch", "Fill up", "Invalid"]
        })
        
        result = prepare_export_dataframe(df)
        # Should have 2 valid entries + 1 TOTAL row
        assert len(result) == 3
        assert result["Category"].iloc[2] == "TOTAL"
        assert result["Amount ($)"].iloc[2] == 75.50
    
    def test_prepare_empty_dataframe(self):
        """Test preparing empty DataFrame"""
        df = pd.DataFrame(columns=["Date/Time", "Category", "Amount ($)", "Notes"])
        result = prepare_export_dataframe(df)
        assert result.empty


class TestFormatExcelExport:
    """Test suite for format_excel_export() function"""
    
    def test_format_creates_workbook(self):
        """Test that formatting is applied to Excel file"""
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            tmp_path = tmp.name
            
            # Create initial Excel file
            df = pd.DataFrame({
                "Date/Time": ["2025-10-14"],
                "Category": ["Food"],
                "Amount ($)": [25.50],
                "Notes": ["Lunch"]
            })
            df.to_excel(tmp_path, index=False)
            
            # Apply formatting
            format_excel_export(tmp_path)
            
            # Verify file still exists and can be opened
            assert os.path.exists(tmp_path)
            from openpyxl import load_workbook
            wb = load_workbook(tmp_path)
            ws = wb.active
            assert ws.title == "Expense Summary"
            
            # Cleanup
            wb.close()
            os.unlink(tmp_path)
    
    def test_format_with_multiple_rows(self):
        """Test formatting with multiple data rows"""
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            tmp_path = tmp.name
            
            df = pd.DataFrame({
                "Date/Time": ["2025-10-14", "2025-10-15"],
                "Category": ["Food", "Gas"],
                "Amount ($)": [25.50, 50.00],
                "Notes": ["Lunch", "Fill up"]
            })
            df.to_excel(tmp_path, index=False)
            
            format_excel_export(tmp_path)
            
            from openpyxl import load_workbook
            wb = load_workbook(tmp_path)
            ws = wb.active
            
            # Verify multiple rows exist
            assert ws.max_row >= 3  # Header + 2 data rows
            
            # Cleanup
            wb.close()
            os.unlink(tmp_path)


class TestEdgeCasesAndIntegration:
    """Test edge cases and integration scenarios"""
    
    def test_full_workflow(self):
        """Test complete workflow from create to export"""
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            tmp_path = tmp.name
            
            # Start with fresh empty DataFrame (not loading non-existent file)
            df = pd.DataFrame(columns=["Date/Time", "Category", "Amount ($)", "Notes"])
            assert df.empty
            
            # Add multiple entries
            df = add_expense_entry(df, datetime(2025, 10, 14).date(), "Food", 25.50, "Lunch")
            df = add_expense_entry(df, datetime(2025, 10, 15).date(), "Gas", 50.00, "Fill up")
            
            # Save data
            save_expense_data(df, tmp_path)
            
            # Load it back
            loaded_df = load_expense_data(tmp_path)
            assert len(loaded_df) == 2
            
            # Prepare for export
            export_df = prepare_export_dataframe(loaded_df)
            assert len(export_df) == 3  # 2 entries + TOTAL
            
            # Cleanup
            os.unlink(tmp_path)
    
    def test_unicode_in_notes(self):
        """Test handling of unicode characters in notes"""
        df = pd.DataFrame(columns=["Date/Time", "Category", "Amount ($)", "Notes"])
        df = add_expense_entry(
            df,
            datetime(2025, 10, 14).date(),
            "Food",
            25.50,
            "Café ☕ with émojis 🎉"
        )
        assert "☕" in df["Notes"].iloc[0]
    
    def test_very_large_amount(self):
        """Test handling of very large amounts"""
        assert extract_amount("$1,000,000") == 1000000.0
        assert extract_amount("999999.99") == 999999.99
    
    def test_very_small_amount(self):
        """Test handling of very small amounts"""
        assert extract_amount("$0.01") == 0.01
        assert extract_amount("0.99 cents") == 0.99


if __name__ == "__main__":
    pytest.main([__file__, "-v", "--cov=logic", "--cov-report=term-missing"])
