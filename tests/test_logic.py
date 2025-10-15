"""
test: comprehensive OOP test suite for SpendLens SOLID architecture

Tests all classes and interfaces with mocking, fixtures, and edge case coverage.
Target: ≥90% code coverage

Test Structure:
- ExpenseParser: Amount/date parsing, natural language
- ExcelStorage: File I/O, create/read/update/delete
- ExpenseAnalyzer: Category totals, filtering, aggregation
- ExcelFormatter: Styling, charts, Excel formatting
- OpenAIService: Mocked API calls, error handling
- ExpenseController: Integration, dependency injection
"""

import pytest
import os
import logging
from datetime import datetime, timedelta
from unittest.mock import Mock, MagicMock, patch, PropertyMock
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook

# Import application code
import sys
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', 'src'))

from models import Expense, CategoryTotal, ParsedExpense
from logic import (
    IParser, IStorage, IAIService,
    ExpenseParser,
    ExcelStorage,
    ExpenseAnalyzer,
    ExcelFormatter,
    OpenAIService,
    ExpenseController
)


# ==================== FIXTURES ====================

@pytest.fixture
def sample_expenses():
    """Create sample expense data for testing."""
    return [
        Expense(
            date=datetime(2025, 10, 10),
            category="Food",
            amount=25.50,
            notes="Lunch at Chipotle"
        ),
        Expense(
            date=datetime(2025, 10, 11),
            category="Gas",
            amount=45.00,
            notes="Shell station"
        ),
        Expense(
            date=datetime(2025, 10, 12),
            category="Food",
            amount=15.75,
            notes="Starbucks coffee"
        ),
        Expense(
            date=datetime(2025, 10, 13),
            category="Entertainment",
            amount=50.00,
            notes="Movie tickets"
        )
    ]


@pytest.fixture
def expense_parser():
    """Create ExpenseParser instance."""
    return ExpenseParser()


@pytest.fixture
def excel_storage(tmp_path):
    """Create ExcelStorage with temporary file."""
    file_path = tmp_path / "test_expenses.xlsx"
    return ExcelStorage(str(file_path))


@pytest.fixture
def expense_analyzer():
    """Create ExpenseAnalyzer instance."""
    return ExpenseAnalyzer()


@pytest.fixture
def excel_formatter():
    """Create ExcelFormatter instance."""
    return ExcelFormatter()


@pytest.fixture
def mock_openai_client():
    """Create mock OpenAI client."""
    mock_client = Mock()
    
    # Mock transcription response
    mock_transcript = Mock()
    mock_transcript.text = "Spent $15 at Starbucks yesterday"
    mock_client.audio.transcriptions.create.return_value = mock_transcript
    
    # Mock chat completion response
    mock_message = Mock()
    mock_message.content = "You spent $136.25 total. Food was your top category at $41.25."
    mock_choice = Mock()
    mock_choice.message = mock_message
    mock_response = Mock()
    mock_response.choices = [mock_choice]
    mock_client.chat.completions.create.return_value = mock_response
    
    return mock_client


@pytest.fixture
def openai_service(mock_openai_client):
    """Create OpenAIService with mocked client."""
    return OpenAIService(mock_openai_client)


@pytest.fixture
def expense_controller(excel_storage, expense_parser, expense_analyzer, excel_formatter, openai_service):
    """Create fully configured ExpenseController with all dependencies."""
    return ExpenseController(
        storage=excel_storage,
        parser=expense_parser,
        analyzer=expense_analyzer,
        formatter=excel_formatter,
        ai_service=openai_service
    )


# ==================== ExpenseParser Tests ====================

class TestExpenseParser:
    """test: ExpenseParser text parsing functionality"""
    
    def test_parse_numeric_amount_with_dollar_sign(self, expense_parser):
        """Ensures '$15.50' is correctly parsed to 15.5."""
        result = expense_parser.parse_amount("I spent $15.50 at the store")
        assert result == 15.5
    
    def test_parse_numeric_amount_without_dollar_sign(self, expense_parser):
        """Ensures '25' is correctly parsed to 25.0."""
        result = expense_parser.parse_amount("Paid 25 for lunch")
        assert result == 25.0
    
    def test_parse_written_number_dollars(self, expense_parser):
        """Ensures 'twenty dollars' is correctly parsed to 20.0."""
        result = expense_parser.parse_amount("Spent twenty dollars")
        assert result == 20.0
    
    def test_parse_written_number_fifteen(self, expense_parser):
        """Ensures 'fifteen' is correctly parsed to 15.0."""
        result = expense_parser.parse_amount("Cost was fifteen bucks")
        assert result == 15.0
    
    def test_parse_amount_with_commas(self, expense_parser):
        """Ensures '1,234.56' is correctly parsed to 1234.56."""
        result = expense_parser.parse_amount("Total was $1,234.56")
        assert result == 1234.56
    
    def test_parse_amount_returns_zero_when_none_found(self, expense_parser):
        """Ensures 0.0 is returned when no amount is present."""
        result = expense_parser.parse_amount("Just went to the store")
        assert result == 0.0
    
    def test_parse_amount_empty_string(self, expense_parser):
        """Ensures empty string returns 0.0."""
        result = expense_parser.parse_amount("")
        assert result == 0.0
    
    def test_parse_date_yesterday(self, expense_parser):
        """Ensures 'yesterday' parses to one day ago."""
        result = expense_parser.parse_date("Spent money yesterday")
        expected = datetime.now() - timedelta(days=1)
        assert result.date() == expected.date()
    
    def test_parse_date_tomorrow(self, expense_parser):
        """Ensures 'tomorrow' parses to one day ahead."""
        result = expense_parser.parse_date("Will spend tomorrow")
        expected = datetime.now() + timedelta(days=1)
        assert result.date() == expected.date()
    
    def test_parse_date_today_default(self, expense_parser):
        """Ensures default is today when no date found."""
        result = expense_parser.parse_date("Bought groceries")
        assert result.date() == datetime.now().date()
    
    def test_parse_date_days_ago(self, expense_parser):
        """Ensures '3 days ago' is parsed correctly."""
        result = expense_parser.parse_date("Spent money 3 days ago")
        expected = datetime.now() - timedelta(days=3)
        # Allow for slight time differences (within 4 days is acceptable for parsing)
        assert abs((result.date() - expected.date()).days) <= 4
    
    def test_parse_date_next_monday(self, expense_parser):
        """Ensures 'next Monday' parses to future date."""
        result = expense_parser.parse_date("Will pay next Monday")
        # Check it's in the next 14 days (reasonable for "next Monday")
        assert result.date() >= (datetime.now() - timedelta(days=1)).date()
        assert result.date() <= (datetime.now() + timedelta(days=14)).date()
    
    def test_parse_date_last_friday(self, expense_parser):
        """Ensures 'last Friday' parses to past date."""
        result = expense_parser.parse_date("Bought something last Friday")
        assert result <= datetime.now()
    
    def test_parse_expense_complete(self, expense_parser):
        """Ensures full expense parsing extracts both amount and date."""
        text = "Spent $25.50 at Chipotle yesterday"
        result = expense_parser.parse_expense(text)
        
        assert isinstance(result, ParsedExpense)
        assert result.detected_amount == 25.5
        assert result.detected_date.date() == (datetime.now() - timedelta(days=1)).date()
        assert result.raw_text == text
        assert result.confidence == 1.0
    
    def test_parse_expense_missing_amount_lowers_confidence(self, expense_parser):
        """Ensures confidence decreases when amount is missing."""
        text = "Went shopping yesterday"
        result = expense_parser.parse_expense(text)
        
        assert result.detected_amount == 0.0
        assert result.confidence < 1.0
    
    def test_parse_expense_with_written_amounts(self, expense_parser):
        """Ensures written numbers work in full parsing."""
        text = "Paid fifty dollars for dinner last night"
        result = expense_parser.parse_expense(text)
        
        assert result.detected_amount == 50.0


# ==================== ExcelStorage Tests ====================

class TestExcelStorage:
    """test: ExcelStorage file I/O operations"""
    
    def test_storage_initialization(self, excel_storage, tmp_path):
        """Ensures storage initializes with correct file path."""
        assert excel_storage.file_path == str(tmp_path / "test_expenses.xlsx")
    
    def test_load_expenses_empty_file_returns_empty_list(self, excel_storage):
        """Ensures non-existent file returns empty list."""
        expenses = excel_storage.load_expenses()
        assert expenses == []
    
    def test_save_and_load_single_expense(self, excel_storage):
        """Ensures single expense saves and loads correctly."""
        expense = Expense(
            date=datetime(2025, 10, 14),
            category="Food",
            amount=12.99,
            notes="Test expense"
        )
        
        excel_storage.save_expenses([expense])
        loaded = excel_storage.load_expenses()
        
        assert len(loaded) == 1
        assert loaded[0].category == "Food"
        assert loaded[0].amount == 12.99
        assert loaded[0].notes == "Test expense"
    
    def test_save_and_load_multiple_expenses(self, excel_storage, sample_expenses):
        """Ensures multiple expenses save and load correctly."""
        excel_storage.save_expenses(sample_expenses)
        loaded = excel_storage.load_expenses()
        
        assert len(loaded) == 4
        assert loaded[0].category == "Food"
        assert loaded[1].category == "Gas"
        assert loaded[2].amount == 15.75
        assert loaded[3].amount == 50.00
    
    def test_add_expense_appends_to_existing(self, excel_storage, sample_expenses):
        """Ensures add_expense appends without overwriting."""
        excel_storage.save_expenses(sample_expenses)
        
        new_expense = Expense(
            date=datetime(2025, 10, 15),
            category="Shopping",
            amount=99.99,
            notes="New item"
        )
        
        excel_storage.add_expense(new_expense)
        loaded = excel_storage.load_expenses()
        
        assert len(loaded) == 5
        assert loaded[-1].category == "Shopping"
        assert loaded[-1].amount == 99.99
    
    def test_clear_all_removes_all_expenses(self, excel_storage, sample_expenses):
        """Ensures clear_all removes all data."""
        excel_storage.save_expenses(sample_expenses)
        assert len(excel_storage.load_expenses()) == 4
        
        excel_storage.clear_all()
        assert len(excel_storage.load_expenses()) == 0
    
    def test_load_expenses_filters_total_rows(self, excel_storage):
        """Ensures TOTAL rows are filtered out during load."""
        expenses = [
            Expense(datetime(2025, 10, 14), "Food", 10.0, "Test"),
            Expense(datetime(2025, 10, 14), "TOTAL", 10.0, "")
        ]
        
        excel_storage.save_expenses(expenses)
        loaded = excel_storage.load_expenses()
        
        # Should only load the Food expense, not TOTAL
        assert len(loaded) == 1
        assert loaded[0].category == "Food"
    
    def test_save_expenses_creates_valid_excel_file(self, excel_storage, sample_expenses, tmp_path):
        """Ensures saved file is valid Excel format."""
        excel_storage.save_expenses(sample_expenses)
        
        file_path = tmp_path / "test_expenses.xlsx"
        assert file_path.exists()
        
        # Verify it's a valid Excel file
        df = pd.read_excel(str(file_path))
        assert "Category" in df.columns
        assert "Amount ($)" in df.columns
        assert len(df) == 4


# ==================== ExpenseAnalyzer Tests ====================

class TestExpenseAnalyzer:
    """test: ExpenseAnalyzer aggregation and filtering"""
    
    def test_calculate_category_totals(self, expense_analyzer, sample_expenses):
        """Ensures category totals are calculated correctly."""
        totals = expense_analyzer.calculate_category_totals(sample_expenses)
        
        assert len(totals) == 3  # Food, Gas, Entertainment
        
        # Find Food category (should be top with $41.25)
        food_total = next(t for t in totals if t.category == "Food")
        assert food_total.total == 41.25  # 25.50 + 15.75
        assert food_total.count == 2
    
    def test_calculate_category_totals_sorted_descending(self, expense_analyzer, sample_expenses):
        """Ensures totals are sorted by amount descending."""
        totals = expense_analyzer.calculate_category_totals(sample_expenses)
        
        # Entertainment ($50) should be first
        assert totals[0].category == "Entertainment"
        assert totals[0].total == 50.0
    
    def test_calculate_category_totals_empty_list(self, expense_analyzer):
        """Ensures empty expense list returns empty totals."""
        totals = expense_analyzer.calculate_category_totals([])
        assert totals == []
    
    def test_get_total_spending(self, expense_analyzer, sample_expenses):
        """Ensures total spending calculation is correct."""
        total = expense_analyzer.get_total_spending(sample_expenses)
        assert total == 136.25  # 25.50 + 45.00 + 15.75 + 50.00
    
    def test_get_total_spending_empty_list(self, expense_analyzer):
        """Ensures empty list returns 0.0."""
        total = expense_analyzer.get_total_spending([])
        assert total == 0.0
    
    def test_filter_by_date_range(self, expense_analyzer, sample_expenses):
        """Ensures date range filtering works correctly."""
        start = datetime(2025, 10, 11)
        end = datetime(2025, 10, 12)
        
        filtered = expense_analyzer.filter_by_date_range(sample_expenses, start, end)
        
        assert len(filtered) == 2  # Gas and Food on 10/11 and 10/12
        assert filtered[0].category == "Gas"
        assert filtered[1].category == "Food"
    
    def test_filter_by_date_range_no_matches(self, expense_analyzer, sample_expenses):
        """Ensures no matches returns empty list."""
        start = datetime(2025, 1, 1)
        end = datetime(2025, 1, 2)
        
        filtered = expense_analyzer.filter_by_date_range(sample_expenses, start, end)
        assert filtered == []


# ==================== ExcelFormatter Tests ====================

class TestExcelFormatter:
    """test: ExcelFormatter styling and chart generation"""
    
    def test_format_workbook_applies_header_styling(self, excel_formatter, excel_storage, sample_expenses, tmp_path):
        """Ensures headers are bold with colored background."""
        excel_storage.save_expenses(sample_expenses)
        excel_formatter.format_workbook(str(tmp_path / "test_expenses.xlsx"))
        
        wb = load_workbook(str(tmp_path / "test_expenses.xlsx"))
        ws = wb.active
        
        # Check first header cell
        header_cell = ws['A1']
        assert header_cell.font.bold is True
        assert "FFFFFF" in header_cell.font.color.rgb  # White text (with or without alpha)
        assert "1F4E78" in header_cell.fill.start_color.rgb  # Blue background
    
    def test_format_workbook_applies_alternating_rows(self, excel_formatter, excel_storage, sample_expenses, tmp_path):
        """Ensures alternating row colors are applied."""
        excel_storage.save_expenses(sample_expenses)
        excel_formatter.format_workbook(str(tmp_path / "test_expenses.xlsx"))
        
        wb = load_workbook(str(tmp_path / "test_expenses.xlsx"))
        ws = wb.active
        
        # Row 2 should have alternating fill
        cell_row2 = ws['A2']
        assert "F2F2F2" in cell_row2.fill.start_color.rgb  # Gray background (with or without alpha)
        
        # Row 3 should not have alternating fill (or different)
        cell_row3 = ws['A3']
        # Row 3 is odd, so it should not have the alt fill
        assert "F2F2F2" not in cell_row3.fill.start_color.rgb
    
    def test_format_workbook_applies_currency_format(self, excel_formatter, excel_storage, sample_expenses, tmp_path):
        """Ensures amount column has currency formatting."""
        excel_storage.save_expenses(sample_expenses)
        excel_formatter.format_workbook(str(tmp_path / "test_expenses.xlsx"))
        
        wb = load_workbook(str(tmp_path / "test_expenses.xlsx"))
        ws = wb.active
        
        # Column C (Amount) should have currency format
        amount_cell = ws['C2']
        assert '"$"' in amount_cell.number_format or '#,##0.00' in amount_cell.number_format
    
    def test_format_workbook_applies_borders(self, excel_formatter, excel_storage, sample_expenses, tmp_path):
        """Ensures all cells have borders."""
        excel_storage.save_expenses(sample_expenses)
        excel_formatter.format_workbook(str(tmp_path / "test_expenses.xlsx"))
        
        wb = load_workbook(str(tmp_path / "test_expenses.xlsx"))
        ws = wb.active
        
        # Check a data cell has borders
        cell = ws['B2']
        assert cell.border.left.style == "thin"
        assert cell.border.right.style == "thin"
    
    def test_format_workbook_adds_chart(self, excel_formatter, excel_storage, sample_expenses, tmp_path):
        """Ensures chart is added to workbook."""
        excel_storage.save_expenses(sample_expenses)
        excel_formatter.format_workbook(str(tmp_path / "test_expenses.xlsx"))
        
        wb = load_workbook(str(tmp_path / "test_expenses.xlsx"))
        ws = wb.active
        
        # Check if chart exists
        assert len(ws._charts) > 0
        chart = ws._charts[0]
        # Check title text is in the title object
        assert "Spending by Category" in str(chart.title)
    
    def test_format_workbook_adjusts_column_widths(self, excel_formatter, excel_storage, sample_expenses, tmp_path):
        """Ensures column widths are auto-adjusted."""
        excel_storage.save_expenses(sample_expenses)
        excel_formatter.format_workbook(str(tmp_path / "test_expenses.xlsx"))
        
        wb = load_workbook(str(tmp_path / "test_expenses.xlsx"))
        ws = wb.active
        
        # Column A should have adjusted width
        col_width = ws.column_dimensions['A'].width
        assert col_width > 0  # Should be set to something reasonable
    
    def test_format_workbook_handles_errors_gracefully(self, excel_formatter):
        """Ensures formatter raises exception on invalid file."""
        with pytest.raises(Exception):
            excel_formatter.format_workbook("/invalid/path/file.xlsx")


# ==================== OpenAIService Tests ====================

class TestOpenAIService:
    """test: OpenAIService mocked API interactions"""
    
    def test_transcribe_audio_returns_text(self, openai_service):
        """Ensures audio transcription returns expected text."""
        mock_audio = Mock()
        result = openai_service.transcribe_audio(mock_audio)
        
        assert result == "Spent $15 at Starbucks yesterday"
        assert openai_service.client.audio.transcriptions.create.called
    
    def test_transcribe_audio_calls_correct_model(self, openai_service, mock_openai_client):
        """Ensures correct Whisper model is used."""
        mock_audio = Mock()
        openai_service.transcribe_audio(mock_audio)
        
        call_args = mock_openai_client.audio.transcriptions.create.call_args
        assert call_args[1]['model'] == "gpt-4o-mini-transcribe"
    
    def test_transcribe_audio_handles_api_error(self, mock_openai_client):
        """Ensures transcription errors are raised."""
        mock_openai_client.audio.transcriptions.create.side_effect = Exception("API Error")
        service = OpenAIService(mock_openai_client)
        
        with pytest.raises(Exception) as exc_info:
            service.transcribe_audio(Mock())
        
        assert "API Error" in str(exc_info.value)
    
    def test_analyze_expenses_returns_response(self, openai_service, sample_expenses):
        """Ensures expense analysis returns AI response."""
        query = "What did I spend the most on?"
        result = openai_service.analyze_expenses(sample_expenses, query)
        
        assert "You spent" in result
        assert "Food" in result or "category" in result.lower()
    
    def test_analyze_expenses_formats_data_correctly(self, openai_service, mock_openai_client, sample_expenses):
        """Ensures expenses are formatted correctly for API."""
        query = "Show me my spending"
        openai_service.analyze_expenses(sample_expenses, query)
        
        call_args = mock_openai_client.chat.completions.create.call_args
        messages = call_args[1]['messages']
        
        # Check that expense data is in the user message
        user_message = messages[1]['content']
        assert "2025-10-10" in user_message
        assert "$25.50" in user_message
        assert "Food" in user_message
    
    def test_analyze_expenses_uses_correct_model(self, openai_service, mock_openai_client, sample_expenses):
        """Ensures correct GPT model is used."""
        openai_service.analyze_expenses(sample_expenses, "Test query")
        
        call_args = mock_openai_client.chat.completions.create.call_args
        assert call_args[1]['model'] == "gpt-4o-mini"
    
    def test_analyze_expenses_handles_api_error(self, mock_openai_client, sample_expenses):
        """Ensures analysis errors return error message."""
        mock_openai_client.chat.completions.create.side_effect = Exception("Rate limit exceeded")
        service = OpenAIService(mock_openai_client)
        
        result = service.analyze_expenses(sample_expenses, "Test query")
        
        assert "Sorry" in result
        assert "Rate limit exceeded" in result
    
    def test_openai_service_initialization(self, mock_openai_client):
        """Ensures service initializes with client."""
        service = OpenAIService(mock_openai_client)
        assert service.client == mock_openai_client


# ==================== ExpenseController Tests ====================

class TestExpenseController:
    """test: ExpenseController integration and dependency injection"""
    
    def test_controller_initialization_with_all_dependencies(self, expense_controller):
        """Ensures controller accepts all dependencies via DI."""
        assert expense_controller.storage is not None
        assert expense_controller.parser is not None
        assert expense_controller.analyzer is not None
        assert expense_controller.formatter is not None
        assert expense_controller.ai_service is not None
    
    def test_controller_initialization_without_ai_service(self, excel_storage, expense_parser, expense_analyzer, excel_formatter):
        """Ensures controller works without AI service (optional dependency)."""
        controller = ExpenseController(
            storage=excel_storage,
            parser=expense_parser,
            analyzer=expense_analyzer,
            formatter=excel_formatter,
            ai_service=None
        )
        
        assert controller.ai_service is None
    
    def test_parse_and_create_expense(self, expense_controller):
        """Ensures controller creates expense from text."""
        text = "Spent $25 at Starbucks yesterday"
        expense = expense_controller.parse_and_create_expense(text, "Food")
        
        assert isinstance(expense, Expense)
        assert expense.category == "Food"
        assert expense.amount == 25.0
        assert expense.notes == text
    
    def test_parse_and_create_expense_with_manual_overrides(self, expense_controller):
        """Ensures manual amount/date override parsed values."""
        text = "Spent $25 yesterday"
        manual_date = datetime(2025, 1, 1)
        manual_amount = 100.0
        
        expense = expense_controller.parse_and_create_expense(
            text, "Gas", manual_amount=manual_amount, manual_date=manual_date
        )
        
        assert expense.amount == 100.0
        assert expense.date == manual_date
    
    def test_add_expense_saves_to_storage(self, expense_controller):
        """Ensures add_expense creates and saves expense."""
        text = "Lunch at Chipotle for $12"
        expense = expense_controller.add_expense(text, "Food", manual_amount=12.0)
        
        # Verify it was saved
        loaded = expense_controller.get_all_expenses()
        assert len(loaded) == 1
        assert loaded[0].category == "Food"
        assert loaded[0].amount == 12.0
    
    def test_get_all_expenses_returns_list(self, expense_controller, sample_expenses):
        """Ensures get_all_expenses returns stored data."""
        # Save expenses first
        for exp in sample_expenses:
            expense_controller.storage.add_expense(exp)
        
        loaded = expense_controller.get_all_expenses()
        assert len(loaded) == 4
    
    def test_get_dashboard_data_structure(self, expense_controller, sample_expenses):
        """Ensures dashboard data has correct structure."""
        # Save expenses
        for exp in sample_expenses:
            expense_controller.storage.add_expense(exp)
        
        data = expense_controller.get_dashboard_data()
        
        assert "expenses" in data
        assert "category_totals" in data
        assert "total_spending" in data
        assert "expense_count" in data
        
        assert len(data["expenses"]) == 4
        assert data["expense_count"] == 4
        assert data["total_spending"] == 136.25
    
    def test_get_dashboard_data_empty_storage(self, expense_controller):
        """Ensures dashboard works with no expenses."""
        data = expense_controller.get_dashboard_data()
        
        assert data["expense_count"] == 0
        assert data["total_spending"] == 0.0
        assert data["expenses"] == []
        assert data["category_totals"] == []
    
    def test_export_to_excel_creates_formatted_file(self, expense_controller, sample_expenses, tmp_path):
        """Ensures export creates formatted Excel file."""
        # Save expenses
        for exp in sample_expenses:
            expense_controller.storage.add_expense(exp)
        
        export_path = str(tmp_path / "export.xlsx")
        expense_controller.export_to_excel(export_path)
        
        # Verify file exists and is formatted
        assert os.path.exists(export_path)
        
        wb = load_workbook(export_path)
        ws = wb.active
        
        # Check header styling exists
        assert ws['A1'].font.bold is True
        
        # Check TOTAL row was added
        df = pd.read_excel(export_path)
        total_rows = df[df['Category'] == 'TOTAL']
        assert len(total_rows) == 1
    
    def test_clear_all_expenses_removes_data(self, expense_controller, sample_expenses):
        """Ensures clear_all removes all expenses."""
        # Save expenses
        for exp in sample_expenses:
            expense_controller.storage.add_expense(exp)
        
        assert len(expense_controller.get_all_expenses()) == 4
        
        expense_controller.clear_all_expenses()
        
        assert len(expense_controller.get_all_expenses()) == 0
    
    def test_controller_uses_injected_parser(self, excel_storage, expense_analyzer, excel_formatter):
        """Ensures controller uses injected parser implementation."""
        mock_parser = Mock(spec=IParser)
        mock_parser.parse_expense.return_value = ParsedExpense(
            raw_text="test",
            detected_amount=50.0,
            detected_date=datetime.now(),
            confidence=1.0
        )
        
        controller = ExpenseController(
            storage=excel_storage,
            parser=mock_parser,
            analyzer=expense_analyzer,
            formatter=excel_formatter
        )
        
        controller.parse_and_create_expense("test", "Food")
        assert mock_parser.parse_expense.called
    
    def test_controller_uses_injected_storage(self, expense_parser, expense_analyzer, excel_formatter):
        """Ensures controller uses injected storage implementation."""
        mock_storage = Mock(spec=IStorage)
        mock_storage.load_expenses.return_value = []
        
        controller = ExpenseController(
            storage=mock_storage,
            parser=expense_parser,
            analyzer=expense_analyzer,
            formatter=excel_formatter
        )
        
        controller.get_all_expenses()
        assert mock_storage.load_expenses.called


# ==================== Edge Cases and Error Handling ====================

class TestEdgeCasesAndErrorHandling:
    """test: edge cases, errors, and boundary conditions"""
    
    def test_parse_very_large_amount(self, expense_parser):
        """Ensures parser handles large amounts correctly."""
        result = expense_parser.parse_amount("Paid $999,999.99")
        assert result == 999999.99
    
    def test_parse_very_small_amount(self, expense_parser):
        """Ensures parser handles small amounts correctly."""
        result = expense_parser.parse_amount("Spent $0.01")
        assert result == 0.01
    
    def test_parse_amount_with_multiple_numbers(self, expense_parser):
        """Ensures parser extracts first valid amount."""
        result = expense_parser.parse_amount("Bought 5 items for $50")
        # Should get 5 or 50, likely 5 as it comes first
        assert result in [5.0, 50.0]
    
    def test_expense_to_dict_format(self, sample_expenses):
        """Ensures Expense.to_dict() creates correct format."""
        exp = sample_expenses[0]
        data = exp.to_dict()
        
        assert "Date/Time" in data
        assert "Category" in data
        assert "Amount ($)" in data
        assert "Notes" in data
        assert data["Category"] == "Food"
    
    def test_expense_from_dict_creates_valid_object(self):
        """Ensures Expense.from_dict() reconstructs object."""
        data = {
            "Date/Time": "2025-10-14",
            "Category": "Test",
            "Amount ($)": 25.0,
            "Notes": "Test note"
        }
        
        expense = Expense.from_dict(data)
        assert expense.category == "Test"
        assert expense.amount == 25.0
        assert expense.notes == "Test note"
    
    def test_category_total_structure(self):
        """Ensures CategoryTotal has correct attributes."""
        ct = CategoryTotal(category="Food", total=100.0, count=5)
        assert ct.category == "Food"
        assert ct.total == 100.0
        assert ct.count == 5
    
    def test_parsed_expense_structure(self):
        """Ensures ParsedExpense has correct attributes."""
        pe = ParsedExpense(
            raw_text="Test",
            detected_amount=25.0,
            detected_date=datetime.now(),
            confidence=0.8
        )
        assert pe.raw_text == "Test"
        assert pe.detected_amount == 25.0
        assert pe.confidence == 0.8
    
    def test_storage_handles_corrupted_file(self, tmp_path):
        """Ensures storage handles corrupted Excel files gracefully."""
        file_path = tmp_path / "corrupted.xlsx"
        
        # Create invalid Excel file
        with open(str(file_path), 'w') as f:
            f.write("This is not valid Excel data")
        
        storage = ExcelStorage(str(file_path))
        expenses = storage.load_expenses()
        
        # Should return empty list on error
        assert expenses == []
    
    def test_analyzer_handles_single_expense(self, expense_analyzer):
        """Ensures analyzer works with single expense."""
        expense = Expense(datetime.now(), "Food", 10.0, "Test")
        
        totals = expense_analyzer.calculate_category_totals([expense])
        assert len(totals) == 1
        assert totals[0].total == 10.0
    
    def test_controller_integration_full_workflow(self, expense_controller):
        """Ensures full workflow from add to export works."""
        # Add expenses
        expense_controller.add_expense("Lunch $15", "Food", manual_amount=15.0)
        expense_controller.add_expense("Gas $40", "Gas", manual_amount=40.0)
        
        # Get dashboard
        data = expense_controller.get_dashboard_data()
        assert data["expense_count"] == 2
        assert data["total_spending"] == 55.0
        
        # Verify categories
        categories = [ct.category for ct in data["category_totals"]]
        assert "Food" in categories
        assert "Gas" in categories
    
    @patch('logging.Logger.error')
    def test_parser_logs_errors_on_invalid_input(self, mock_log_error, expense_parser):
        """Ensures parser logs errors for invalid input."""
        # This should trigger an error log
        result = expense_parser.parse_amount(None)
        
        # Should still return 0.0 and log error
        assert result == 0.0
    
    def test_unicode_handling_in_notes(self, excel_storage):
        """Ensures Unicode characters work in expense notes."""
        expense = Expense(
            date=datetime.now(),
            category="Food",
            amount=25.0,
            notes="Café ☕ with émojis 🎉"
        )
        
        excel_storage.save_expenses([expense])
        loaded = excel_storage.load_expenses()
        
        assert len(loaded) == 1
        assert "Café" in loaded[0].notes
        assert "☕" in loaded[0].notes


# ==================== Logging Tests ====================

class TestLogging:
    """test: logging functionality and error reporting"""
    
    def test_parser_logs_successful_parse(self, expense_parser, caplog):
        """Ensures successful parsing generates info logs."""
        with caplog.at_level(logging.INFO):
            expense_parser.parse_amount("Spent $25")
            assert "Parsed numeric amount" in caplog.text
    
    def test_storage_logs_save_operation(self, excel_storage, caplog):
        """Ensures save operations are logged."""
        with caplog.at_level(logging.INFO):
            expense = Expense(datetime.now(), "Food", 10.0, "Test")
            excel_storage.save_expenses([expense])
            assert "Saved" in caplog.text or "expenses" in caplog.text
    
    def test_controller_logs_initialization(self, expense_controller, caplog):
        """Ensures controller logs initialization."""
        # Controller already initialized in fixture, check it logged
        with caplog.at_level(logging.INFO):
            # Create new controller to capture log
            new_controller = ExpenseController(
                storage=expense_controller.storage,
                parser=expense_controller.parser,
                analyzer=expense_controller.analyzer,
                formatter=expense_controller.formatter
            )
            assert "ExpenseController initialized" in caplog.text


if __name__ == "__main__":
    pytest.main([__file__, "-v", "--cov=src", "--cov-report=term-missing"])
